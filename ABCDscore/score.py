import os
from urllib.parse import urlparse
import requests
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import csv

wb = load_workbook("C:/Users/arvid/Documents/xcel/test1.xlsx")
ws = wb.get_sheet_by_name('Sheet1')

#wordlist = load_workbook("c:/Users/arvid/Documents/GitHub/DEPI/WebsiteWordsScaper/wordlist_scrapper/wordlist.csv")

wbwrite = load_workbook("C:/Users/arvid/Documents/xcel/test.xlsx")
wbwritesheet = wbwrite.active

r = 1
c = 1 

n_wordlist = ["energiebron", "energie vermindering", "energie reductie", "energie-intensiteit", "energiegebruik", "energieverbruik",
              "waterverbruik", "waterbron", "wateronttrekking", "waterafvoer", "watergebruik", "afvalwater", "grondwater",
              "broeikasgas", "CO2", "CO²",
              "emissie", "uitstoot", "vervuiling", "zure regen", "uitstoot", "fijnstof", "fijn stof", "vervuilende stof", "filtertechniek", "luchtzuiverheid", "zuiveringstechnologie",
              "impact", "milieu-impact", "impact op het milieu", "milieu impact", "milieu", "mobiliteit", "vervoer", "verplaatsing", "fiets", "auto", "staanplaatsen", "parking", "openbaar vervoer", "klimaatimpact", "impact op het klimaat", "klimaatsverandering", "green deal",
              "gezondheid", "reclyclage", "recycleren", "biodiversiteit", "afval", "afvalproductie", "vervuiling",
              "klimaat", "klimaatsverandering", "klimaatopwarming", "opwarming", "scope",
              "milieubeleid", "hernieuwbare energie", "verspilling", "milieucriteria", "planeet", "klimaatsbeleid", "milieunormen",
              "klimaatactie", "leven in het water", 'leven op het land',
              "duurzaamheidsdoelstellingen", "ontwikkelingsdoelen", "ontwikkelingsdoelstelling", "duurzaamheidsrapportering"]





for row in ws.rows:
    n=0
    m=0
    d = dict()
    a = []
    URL = row[0].value
    print('#########################################################')
    print(URL)
    domain = urlparse(URL).netloc
    print(domain)
    print('#########################################################')
    os.system(
        f"cd C:/Users/arvid/Documents/GitHub/DEPI/WebsiteWordsScaper/wordlist_scrapper && scrapy crawl webcrawler -a start_url={URL} -a domains={domain} > wordlist.csv")

    with open('C:/Users/arvid/Documents/GitHub/DEPI/WebsiteWordsScaper/wordlist_scrapper/wordlist.csv', 'r') as csv_file:
        csv_reader = csv.reader(csv_file)

        for line in csv_reader:

            for word in line:
                if word in d:
                    d[word] = d[word] + 1
                else:
                    d[word] = 1

        for key in list(d.keys()):
            if key in n_wordlist:
                n = n+1
            else:
                m = m+1

        n_perc = round(n/70, 2)
        m_perc = round(m/70,2)
        print("natuurlijk:" + str(n) + "  perc:" + str(n_perc))
        print("menselijk:" + str(m) + "  perc:" + str(m_perc))
        
        
        wbwritesheet.cell(row=r, column=c).value = n_perc
        wbwritesheet.cell(row=r, column=c+1).value = m_perc
        wbwrite.save("C:/Users/arvid/Documents/xcel/test.xlsx")
    
    os.remove(
        r'WebsiteWordsScaper\wordlist_scrapper\wordlist.csv')
    
    r = r+1
    
    