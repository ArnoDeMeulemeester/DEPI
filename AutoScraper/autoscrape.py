from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook


wb = load_workbook("C:/Users/arvid/Documents/xcel/workfile.xlsx")
ws = wb.get_sheet_by_name('Sheet1')
wbwrite = load_workbook("C:/Users/arvid/Documents/xcel/test.xlsx")
wbwritesheet = wbwrite.active
r = 1
c = 1

for row in ws.rows:

    PATH = "C:\Program Files (x86)\chromedriver.exe"
    driver = webdriver.Chrome(PATH)

    print(row[1].value)
    naam = row[1].value
    naam = naam.replace(" ", "")
    print(naam)

    driver.get(
        "https://duckduckgo.com/")

    try:
        element1 = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.ID, "search_form_input_homepage"))
        )
        search = driver.find_element_by_id(
            "search_form_input_homepage")
        search.send_keys(naam + ' belgie website')
        search.send_keys(Keys.ENTER)
    finally:
        print("Onderneming opzoeken was succesvol")
        time.sleep(1)

    try:
        link = driver.find_element_by_class_name(
            "result__url__domain")
        print(link.text)

        wbwritesheet.cell(row=r, column=c).value = link.text
        wbwrite.save("C:/Users/arvid/Documents/xcel/test.xlsx")

    finally:
        time.sleep(1)
        driver.quit()
        r += 1
