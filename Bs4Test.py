# Testscenario van de tutorial

from ast import Try
from logging import exception
from operator import truediv
from re import T
from time import sleep
import requests
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# wb = load_workbook("C:/Users/arvid/Documents/xcel/workfile.xlsx")
wb = load_workbook("C:/Users/arnod/Documents/HoGent/wb.xlsx")
ws = wb.get_sheet_by_name('Sheet1')
# wbwrite = load_workbook("C:/Users/arvid/Documents/xcel/test.xlsx")
wbwrite = load_workbook("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
wbwritesheet = wbwrite.active
r = 1
c = 1

def isNumber(x):
  if x.isdigit():
    return True
  else:
    return False

for row in ws.rows:
    naam = row[1].value
    print(f"{r}. {naam}")
    naam = naam.replace(" ", "")

# url site
    try:
      if(isNumber(naam)):
        raise Exception("De naam van de kmo is een getal")
      else:
        url = 'https://www.bing.com/search?q=' + naam + "+website+Belgie"
        res = requests.get(url)
        # Connect to URL
        html = requests.get(url, headers={"User-Agent": "Requests"})
        obj = BeautifulSoup(html.text, "html.parser")
        link = obj.findAll('cite')[0].get_text()
        print(link)
        print("--------------------------------------")

        wbwritesheet.cell(row=r, column=c).value = link
        # wbwrite.save("C:/Users/arvid/Documents/xcel/test.xlsx")
        wbwrite.save("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
        r += 1
    except Exception as e:
      print(e)
      wbwritesheet.cell(row=r, column=c).value = "website niet gevonden"
      wbwrite.save("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
      r += 1