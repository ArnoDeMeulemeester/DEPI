from io import StringIO
from functools import partial
from scrapy.http import Request
#from scrapy.spiders import BaseSpider
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor
from scrapy.item import Item

import requests
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
# TODO Replace "allowed_domains", "start_urls" and "wordlist" with your own data.

wb = load_workbook("C:/Users/arvid/Documents/xcel/test1.xlsx")
#wb = load_workbook("C:/Users/arnod/Documents/HoGent/wb.xlsx")
ws = wb.get_sheet_by_name('Sheet1')
wbwrite = load_workbook("C:/Users/arvid/Documents/xcel/test.xlsx")
wbwritesheet = wbwrite.active


def find_all_substrings(string, sub):

    import re
    starts = [match.start() for match in re.finditer(re.escape(sub), string)]
    return starts


class WebsiteSpider(CrawlSpider):

    name = "webcrawler"
    allowed_domains = []
    #start_urls = ["https://www.hedinautomotive.be/"]
    rules = [Rule(LinkExtractor(), follow=True, callback="check_buzzwords")]

    crawl_count = 0
    words_found = 0

    def __init__(self, *args, **kwargs):
        super(WebsiteSpider, self).__init__(*args, **kwargs)

        self.start_urls = [kwargs.get('start_url')]
        self.allowed_domains = [kwargs.get('domains')]

    def check_buzzwords(self, response):

        self.__class__.crawl_count += 1

        crawl_count = self.__class__.crawl_count

        wordlist = [
            "energiebron", "energie vermindering", "energie reductie", "energie-intensiteit", "energiegebruik", "energieverbruik",
            "waterverbruik", "waterbron", "wateronttrekking", "waterafvoer", "watergebruik", "afvalwater", "grondwater",
            "broeikasgas", "CO2", "CO²", "CO2",
            "emissie", "uitstoot", "vervuiling", "zure regen", "uitstoot", "fijnstof", "fijn stof", "vervuilende stof", "filtertechniek", "luchtzuiverheid", "zuiveringstechnologie",
            "impact", "milieu-impact", "impact op het milieu", "milieu impact", "milieu", "mobiliteit", "vervoer", "verplaatsing", "fiets", "auto", "staanplaatsen", "parking", "openbaar vervoer", "klimaatimpact", "impact op het klimaat", "klimaatsverandering", "green deal",
            "gezondheid", "reclyclage", "recycleren", "biodiversiteit", "afval", "afvalproductie", "vervuiling",
            "klimaat", "klimaatsverandering", "klimaatopwarming", "opwarming", "scope",
            "milieubeleid", "hernieuwbare energie", "verspilling", "milieucriteria", "planeet", "klimaatsbeleid", "milieunormen",
            "schoon water en sanitair", "betaalbare en duurzame energie", "duurzame steden en gemeenschappen", "verantwoorde consumptie en productie", "klimaatactie", "leven in het water", 'leven op het land',
            "duurzaamheidsdoelstellingen", "ontwikkelingsdoelen", "ontwikkelingsdoelstelling", "duurzaamheidsrapportering", "geslacht", "gendergelijkheid", "man/vrouw verhouding", "ratio man/vrouw",
            "salaris man/vrouw", "discriminatie", "genderneutraal", "sociale relaties", "werkvloer", "solidair gedrag", "betrokkenheid van werknemers", "rekrutering", "rekruteringstijd", 
            "arbeidsovereenkomst", "werknemer", "werknemers", "diversiteitsbeleid",  "loopbaan", "carriere", "groeiopportuniteiten", "geslacht", "gendergelijkheid", "man/vrouw verhouding", 
            "ratio man/vrouw", "salaris man/vrouw", "discriminatie", "genderneutraal", "duurzaamheidscommissie", "rechten van werknemers", "arbeisrechten", "rechten en plichten van werknemers", 
            "arbeidsomstandigheden", "algemene rechten en plichten", "mensenrechten", "recht op vrijheid", "sociale relaties", "werkvloer", "solidair gedrag",  "betrokkenheid van werknemers", 
            "rekrutering", "rekruteringsbeleid", "rekruteringsverloop", "rekruteringstijd", "arbeidsovereenkomst", "werknemer", "werknemers", "diversiteitsbeleid", "loopbaan", "carrière", 
            "carrièreontwikkeling", "groei opportuniteiten", "groeikansen", "doorstroommogelijkheden", "promotie", "demografisch", "personeelsbestand", "promotie", "human resources", "HR", 
            "personeelsbeleid", "vergoeding", "beloning", "bonus", "stabiliteit van werknemers", "stabiliteit", "bedrijfscultuur", "loyaliteit van werknemers", "personeelsbehoud", 
            "retentie personeel", "loyaliteitsbonus", "personeelsverloop", "leeftijdsstructuur", "afwezigheid", "afwezigheidsratio", "tevredenheid op het werk", "opvolgingsbeheer", 
            "prestatiebeleid", "prestatie", "ziekte", "verzuim", "aanwezigheid", "aanwezigheidsratio", "preventie", "pesterij", "ongewenst gedrag", "klacht", "gezondheid van werknemers", 
            "welzijn van werknemers", "managers", "arbeiders", "bedienden", "medewerkers", "incidenten op het werk", "incidenten", "discriminatie", "gezondheid en veiligheid op het werk", 
            "intimidatie", "intimiderend gedrag", "vakbond", "opleidingsbeleid", "training", "opleiding", "vaardigheden van werknemers", "kennis van werknemers", "werknemersvaardigheden", 
            "competenties van werknemer", "werknemercompetenties", "talent", "vakbekwaamheid", "kinderarbeid", "goede gezondheid en welzijn", "gender-gelijkheid", 
            "waardig werk en economische groei", "ongelijkheid verminderen", "vrede", "veiligheid en sterke publieke diensten", "duurzaamheidsdoelstelling(en)", "ontwikkelingsdoelen", 
            "ontwikkelingsdoelstelling", "duurzaamheidsrapportering"
        ]

        url = response.url
        contenttype = response.headers.get(
            "content-type", "").decode('utf-8').lower()
        data = response.body.decode('utf-8')

        listw = []
        for word in wordlist:
            substrings = find_all_substrings(data, word)
            for pos in substrings:
                ok = False
                if not ok:
                    self.__class__.words_found += 1
                    listw.append(word)
        print(*listw, sep=',')
        return Item()

    def _requests_to_follow(self, response):
        if getattr(response, "encoding", None) != None:
            return CrawlSpider._requests_to_follow(self, response)
        else:
            return []