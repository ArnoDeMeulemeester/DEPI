from openpyxl import Workbook, load_workbook
import time
import requests
from bs4 import BeautifulSoup
import operator
from collections.abc import Counter
import pandas as pd

#read excel file
df =  pd.read_excel('/Users/lauraingelbrecht/Downloads/toegepaste informatica jaar 2/Semester 2/Data Engineering Project I/Map1.xlsx')
#print(df) 
wb = Workbook()
wb = load_workbook("/Users/lauraingelbrecht/Downloads/toegepaste informatica jaar 2/Semester 2/Data Engineering Project I/Map1.xlsx")
ws = wb.sheetnames('Sheet1')


for row in ws.rows:
    print(row[0].value)
    time.sleep(0.25)

producten_alle_paginas = []

for i in range(1,15):
    response = requests.get(f"") #todo url moet er in komen
    content = response.content
    parser = BeautifulSoup(content, 'html.parser')
    body = parser.body
    producten = body.find_all(class_="p")
    producten_alle_paginas.extend(producten)
    
len(producten_alle_paginas)

#meest voorkomende woorden in een website

def start(url):
    worldlist = []
    source_code = requests.get(url).text
    soup = BeautifulSoup(source_code, 'html.parser')
    for each_text in soup.findAll('div', {'class': 'panel-pane pane-top-news-new no-title block'}):
        content = each_text.text
        words = content.lower().split()
        for each_word in words:
            worldlist.append(each_word)
        clean_wordlist(worldlist)

def clean_wordlist(wordlist):
    clean_list = []
    for word in wordlist:
        symbols = "!@#$%^&*()_-+={[}]|\;:\"<>?/., "
        for i in range(len(symbols)):
            word = word.replace(symbols[i], '')
        if len(word) > 0:
            clean_list.append(word)
    create_dict(clean_list)

def create_dict(clean_list):
    word_count = {}
    for word in clean_list:
        if word in word_count:
            word_count[word] += 1
        else:
            word_count[word] = 1
    c = Counter(word_count)
    top = c.most_common(20)
    print(top)

if __name__ != '__main__':
        pass
else:
    url = "https://itsmycode.com/importerror-no-module-named-requests/"
    start(url)

