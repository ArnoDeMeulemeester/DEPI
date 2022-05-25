from openpyxl import Workbook, load_workbook

wb = load_workbook("C:/Users/arnod/Documents/HoGent/wb.xlsx")
ws = wb.get_sheet_by_name('Sheet1')

wbwrite = load_workbook("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
wbwritesheet = wbwrite.active

c=1
wetenschappelijkeEnTechnischeActiviteiten = ["wetenschap", "technische activiteiten"]
handel = ["handel", "groothandel", "detailhandel"]
bouwnijverheid = ["bouwnijverheid", "bouw", "bouwen", "slopen", "schrijnwerk", "gebouwen"]
administratieveEnOndersteunendeDiensten = ["administratie", "ondersteunen"]
horeca = ["hotel", "restaurant", "café", "bedienen", "catering"]
informatieEnCommunicatie = ["informatie", "communicatie"]
industrie = ["industrie", "industrieel"]

def bepaalSector(omschrijving):
  lijst = omschrijving.split()
  lowerCaseLijst = []
  for word in lijst:
    lowerCaseLijst.append(word.lower())
    
  print(lowerCaseLijst)
  
  for x in lowerCaseLijst: 
    if any(x in s for s in horeca):
      sector = "horeca"
      break
    elif any(x in s for s in industrie):
      sector = "industrie"
      break
    elif any(x in s for s in wetenschappelijkeEnTechnischeActiviteiten):
      sector = "wetenschappelijke en technische activiteiten"
      break
    elif any(x in s for s in bouwnijverheid):
      sector = "bouwnijverheid"
      break
    elif any(x in s for s in administratieveEnOndersteunendeDiensten):
      sector = "administratieve en ondersteunende diensten"
      break
    elif any(x in s for s in informatieEnCommunicatie):
      sector = "informatie en communicatie"
      break
    elif any(x in s for s in handel):
      sector = "handel"
      break
    else:
      sector = "overige diensten"
  
  return sector

for row in ws.rows:
    naam = row[1].value
    r = int(row[0].value)
    print(f"{r}. {naam}")

    try:
      omschrijving = row[2].value
      print(omschrijving)
      sector = bepaalSector(omschrijving)
      print(sector)
      print("--------------------------------------")
      wbwritesheet.cell(row=r, column=c).value = sector
      # wbwrite.save("C:/Users/arvid/Documents/xcel/test.xlsx")
      wbwrite.save("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
      r += 1
    except Exception as e:
      print(e)
      wbwrite.save("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
      r += 1