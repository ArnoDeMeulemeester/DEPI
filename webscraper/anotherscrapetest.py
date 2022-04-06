
import requests
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

wb = load_workbook("C:/Users/arvid/Documents/xcel/test1.xlsx")
ws = wb.get_sheet_by_name('Sheet1')
wbwrite = load_workbook("C:/Users/arvid/Documents/xcel/test.xlsx")
wbwritesheet = wbwrite.active
r = 1
c = 1

for row in ws.rows:
    naam = row[0].value
    naam = naam.replace(" ", "+")
    print(naam)

    response = requests.get(
        f'https://autocomplete.clearbit.com/v1/companies/suggest?query={naam}')
    data = response.json()
    
    if len(data) > 0:
        domain = data[0]['domain']
        print(domain)

    print("--------------------------------------")

    wbwritesheet.cell(row=r, column=c).value = domain
    wbwrite.save("C:/Users/arvid/Documents/xcel/test.xlsx")

    r += 1
