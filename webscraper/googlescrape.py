from googlesearch import search
import requests
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time

wb = load_workbook("C:/Users/arnod/Documents/HoGent/wb.xlsx")
# wb = load_workbook("C:/Users/arnod/Documents/xcel/workfile.xlsx")
ws = wb.get_sheet_by_name('Sheet1')
wbwrite = load_workbook("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")
# wbwrite = load_workbook("C:/Users/arvid/Documents/xcel/test.xlsx")
wbwritesheet = wbwrite.active

r = 1
c = 1

for row in ws.rows:
    time.sleep(2)
    naam = row[1].value
    naam = naam.replace(" ", "+")
    print(naam)
    links = search(naam+" website Belgie",  tld='com', lang='en', num=1, start=0, stop=1, pause=4.0)
    links = list(links)
    domain = links[0]
    print(domain)
    print("--------------------------------------")

    wbwritesheet.cell(row=r, column=c).value = domain
    # wbwrite.save("C:/Users/arvid/Documents/xcel/test.xlsx")
    wbwrite.save("C:/Users/arnod/Documents/HoGent/wbwrite.xlsx")

    r += 1
